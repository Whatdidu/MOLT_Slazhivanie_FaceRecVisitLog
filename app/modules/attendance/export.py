"""
Модуль экспорта данных посещаемости.

Поддерживаемые форматы:
- JSON
- Excel (xlsx)
"""

import io
import json
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.modules.attendance.models import AttendanceLogResponse, AttendanceStatsResponse


class AttendanceExporter:
    """Экспортёр данных посещаемости."""

    # ============== JSON Export ==============

    @staticmethod
    def to_json(
        logs: list[AttendanceLogResponse],
        pretty: bool = True,
    ) -> str:
        """
        Экспорт в JSON строку.

        Args:
            logs: Список записей посещений
            pretty: Форматировать с отступами

        Returns:
            JSON строка
        """
        data = [
            {
                "id": log.id,
                "employee_id": log.employee_id,
                "employee_name": log.employee_name,
                "event_type": log.event_type.value,
                "timestamp": log.timestamp.isoformat(),
                "confidence": log.confidence,
                "trace_id": log.trace_id,
            }
            for log in logs
        ]

        if pretty:
            return json.dumps(data, ensure_ascii=False, indent=2)
        return json.dumps(data, ensure_ascii=False)

    @staticmethod
    def to_json_bytes(
        logs: list[AttendanceLogResponse],
        pretty: bool = True,
    ) -> bytes:
        """Экспорт в JSON как bytes для скачивания."""
        return AttendanceExporter.to_json(logs, pretty).encode("utf-8")

    # ============== Excel Export ==============

    @staticmethod
    def to_excel(
        logs: list[AttendanceLogResponse],
        title: str = "Журнал посещений",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """
        Экспорт в Excel (xlsx).

        Args:
            logs: Список записей посещений
            title: Заголовок отчёта
            start_date: Начало периода (для заголовка)
            end_date: Конец периода (для заголовка)

        Returns:
            Bytes содержимого xlsx файла
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Посещения"

        # Стили
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")

        # Заголовок отчёта
        ws.merge_cells("A1:G1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = center_align

        # Период
        if start_date and end_date:
            ws.merge_cells("A2:G2")
            ws["A2"] = f"Период: {start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"
            ws["A2"].alignment = center_align

        # Дата формирования
        ws.merge_cells("A3:G3")
        ws["A3"] = f"Сформировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws["A3"].alignment = center_align

        # Заголовки таблицы
        headers = ["№", "ID", "Сотрудник", "Событие", "Дата и время", "Уверенность", "Trace ID"]
        header_row = 5

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        # Данные
        event_type_map = {"entry": "Вход", "exit": "Выход"}

        for idx, log in enumerate(logs, 1):
            row = header_row + idx
            ws.cell(row=row, column=1, value=idx).border = border
            ws.cell(row=row, column=2, value=log.employee_id).border = border
            ws.cell(row=row, column=3, value=log.employee_name or f"#{log.employee_id}").border = border
            ws.cell(row=row, column=4, value=event_type_map.get(log.event_type.value, log.event_type.value)).border = border
            ws.cell(row=row, column=5, value=log.timestamp.strftime("%d.%m.%Y %H:%M:%S")).border = border
            ws.cell(row=row, column=6, value=f"{log.confidence:.1%}").border = border
            ws.cell(row=row, column=7, value=log.trace_id).border = border

            # Выравнивание
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=2).alignment = center_align
            ws.cell(row=row, column=4).alignment = center_align
            ws.cell(row=row, column=5).alignment = center_align
            ws.cell(row=row, column=6).alignment = center_align

        # Ширина колонок
        column_widths = [5, 8, 25, 10, 20, 12, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Итого
        total_row = header_row + len(logs) + 2
        ws.cell(row=total_row, column=1, value="Итого записей:")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=len(logs))

        # Сохраняем в bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ============== Stats Export ==============

    @staticmethod
    def stats_to_excel(
        stats_list: list[AttendanceStatsResponse],
        title: str = "Статистика посещаемости",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """
        Экспорт статистики в Excel.

        Args:
            stats_list: Список статистики по сотрудникам
            title: Заголовок отчёта
            start_date: Начало периода
            end_date: Конец периода

        Returns:
            Bytes содержимого xlsx файла
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика"

        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")

        # Заголовок
        ws.merge_cells("A1:F1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = center_align

        if start_date and end_date:
            ws.merge_cells("A2:F2")
            ws["A2"] = f"Период: {start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"
            ws["A2"].alignment = center_align

        # Заголовки таблицы
        headers = ["ID", "Сотрудник", "Дней", "Часов", "Ср. приход", "Ср. уход"]
        header_row = 4

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        # Данные
        for idx, stats in enumerate(stats_list, 1):
            row = header_row + idx
            ws.cell(row=row, column=1, value=stats.employee_id).border = border
            ws.cell(row=row, column=2, value=stats.employee_name).border = border
            ws.cell(row=row, column=3, value=stats.total_days).border = border
            ws.cell(row=row, column=4, value=f"{stats.total_hours:.1f}").border = border
            ws.cell(row=row, column=5, value=stats.avg_arrival_time or "-").border = border
            ws.cell(row=row, column=6, value=stats.avg_departure_time or "-").border = border

            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=3).alignment = center_align
            ws.cell(row=row, column=4).alignment = center_align
            ws.cell(row=row, column=5).alignment = center_align
            ws.cell(row=row, column=6).alignment = center_align

        # Ширина колонок
        column_widths = [8, 25, 8, 10, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Сохраняем
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


# Удобные функции-обёртки
def export_to_json(logs: list[AttendanceLogResponse]) -> bytes:
    """Быстрый экспорт в JSON."""
    return AttendanceExporter.to_json_bytes(logs)


def export_to_excel(
    logs: list[AttendanceLogResponse],
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
) -> bytes:
    """Быстрый экспорт в Excel."""
    return AttendanceExporter.to_excel(logs, start_date=start_date, end_date=end_date)
